"""LapSimParser — read and validate LapSim Excel exports.

Handles header row detection (LapSim files often have metadata in the first
few rows) and validates that the data is clean (single participant, single
exercise).
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import openpyxl
from app.utils.logger import get_logger
from app.utils.config import LC_MIN_SESSIONS

logger = get_logger(__name__)


@dataclass
class TrialRecord:
    """Performance metrics for a single simulator attempt."""
    trial_number: int       # 1-based chronological sequence
    start_time: str         # ISO 8601 string
    total_time_s: Optional[float]
    score: Optional[float]
    tissue_damage: Optional[int]
    status: Optional[str]   # "Pass" / "Failed"
    left_instrument_time_s: Optional[float] = None
    right_instrument_time_s: Optional[float] = None


@dataclass
class ParsedDataset:
    """Validated dataset ready for model fitting."""
    participant: str       # Login
    participant_name: str  # Full name if available
    exercise: str
    course: str
    source_file: str
    trials: List[TrialRecord]
    warnings: List[str]


class LapSimParser:
    """Parser for .xlsx exports from Surgical Science LapSim simulators."""

    def list_sheets(self, path: str) -> List[str]:
        """Return all sheet names in the Excel file."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            return wb.sheetnames
        except Exception as e:
            logger.error("Failed to list sheets in %s: %s", path, e)
            return []

    def get_participants(self, path: str, sheet_name: str) -> List[Dict[str, str]]:
        """Return a list of unique participants who have the minimum required sessions.
        
        Returns:
            List of dicts with 'login', 'firstname', 'lastname' keys.
        """
        header_row_idx = self._detect_header_row(path, sheet_name)
        if header_row_idx is None:
            return []
            
        df = pd.read_excel(path, sheet_name=sheet_name, skiprows=header_row_idx)
        
        # We need at least Login and Start Time to count valid sessions
        if "Login" not in df.columns or "Start Time" not in df.columns:
            return []
            
        # Drop rows where critical data is missing (these don't count as valid sessions)
        df = df.dropna(subset=["Login", "Start Time"])
        
        # Filter by Session Count
        session_counts = df.groupby("Login").size()
        
        # Keep only logins that meet or exceed the global minimum requirement
        valid_logins = session_counts[session_counts >= LC_MIN_SESSIONS].index
        
        # Filter the dataframe to only include these valid participants
        df_valid = df[df["Login"].isin(valid_logins)]

        if df_valid.empty:
            return []
        
        # Determine available name columns
        cols = ["Login"]
        if "Firstname" in df_valid.columns: cols.append("Firstname")
        if "Lastname" in df_valid.columns: cols.append("Lastname")
        
        unique_participants = df_valid[cols].drop_duplicates()
        
        results = []
        for _, row in unique_participants.iterrows():
            # Safely extract names and ignore pandas "nan" strings
            fname = str(row.get("Firstname", "")).strip()
            lname = str(row.get("Lastname", "")).strip()
            
            if fname.lower() == "nan": fname = ""
            if lname.lower() == "nan": lname = ""

            results.append({
                "login": str(row["Login"]),
                "firstname": fname,
                "lastname": lname
            })
        return results

    def get_data_row_count(self, path: str, sheet_name: str) -> int:
        """Return the number of data rows in a sheet (excluding header)."""
        header_idx = self._detect_header_row(path, sheet_name)
        if header_idx is None:
            return 0
            
        try:
            # Read only the 'Login' column to be fast
            df = pd.read_excel(path, sheet_name=sheet_name, skiprows=header_idx, usecols=["Login"])
            return len(df.dropna(subset=["Login"]))
        except Exception:
            return 0

    def parse(self, path: str, sheet_name: str, login: Optional[str] = None) -> ParsedDataset:
        """Read a specific sheet and return a validated dataset.

        If login is provided, filters data to that participant.
        If login is None and multiple participants exist, raises ValueError.

        Raises:
            ValueError: If the sheet does not contain valid data or participant is ambiguous.
            FileNotFoundError: If the path is invalid.
        """
        # 1. Detect the real header row
        header_row_idx = self._detect_header_row(path, sheet_name)
        if header_row_idx is None:
            raise ValueError(
                f"Sheet '{sheet_name}' is not a valid LapSim export (no 'Login' column found)."
            )

        # 2. Load with pandas
        df = pd.read_excel(path, sheet_name=sheet_name, skiprows=header_row_idx)

        # 3. Validation: Required columns
        required = ["Login", "Start Time", "Task Name"]
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise ValueError(f"Missing required LapSim columns: {', '.join(missing)}")

        # 4. Filter by Login
        valid_rows = df.dropna(subset=["Login", "Start Time", "Task Name"])
        if login:
            valid_rows = valid_rows[valid_rows["Login"].astype(str) == login]
            if valid_rows.empty:
                raise ValueError(f"No data found for participant '{login}'.")
        else:
            unique_logins = valid_rows["Login"].unique()
            if len(unique_logins) == 0:
                raise ValueError(f"No participant data found in sheet '{sheet_name}'.")
            if len(unique_logins) > 1:
                names = ", ".join(map(str, unique_logins))
                raise ValueError(
                    f"Multiple participants found: {names}. "
                    "Select a participant to continue."
                )

        # 5. Validation: Single exercise
        unique_tasks = valid_rows["Task Name"].unique()
        if len(unique_tasks) > 1:
            names = ", ".join(map(str, unique_tasks))
            raise ValueError(
                f"Multiple exercises found for this participant: {names}. "
                "The parser requires a single exercise per analysis."
            )

        # 6. Non-fatal Warnings
        warnings = []
        if "Total Time (s)" not in df.columns:
            warnings.append("'Total Time (s)' column missing; time-based metric unavailable.")
        
        if len(valid_rows) < LC_MIN_SESSIONS:
            warnings.append(
                f"Fewer than {LC_MIN_SESSIONS} trials ({len(valid_rows)}); "
                "model fit may be unstable."
            )
        
        if "Status" in df.columns and (valid_rows["Status"] == "Failed").any():
            warnings.append("Failed attempts detected in dataset; included in analysis.")

        # 7. Chronological Sorting
        valid_rows = valid_rows.copy()
        valid_rows["Start Time"] = pd.to_datetime(valid_rows["Start Time"])
        valid_rows = valid_rows.sort_values("Start Time")

        trials = []
        for i, (_, row) in enumerate(valid_rows.iterrows(), 1):
            left_t = self._get_float(row, "Left Instrument Time (s)")
            right_t = self._get_float(row, "Right Instrument Time (s)")
            total_t = self._get_float(row, "Total Time (s)")
            # Grasping sheet has no "Total Time (s)" — derive from instrument times
            if total_t is None and left_t is not None and right_t is not None:
                total_t = max(left_t, right_t)
            elif total_t is None and (left_t is not None or right_t is not None):
                total_t = left_t or right_t

            trials.append(TrialRecord(
                trial_number=i,
                start_time=str(row["Start Time"]),
                total_time_s=total_t,
                score=self._get_float(row, "Score"),
                tissue_damage=self._get_int(row, "Tissue Damage (#)"),
                status=str(row["Status"]) if "Status" in df.columns and pd.notnull(row["Status"]) else None,
                left_instrument_time_s=left_t,
                right_instrument_time_s=right_t,
            ))

        final_login = login if login else str(valid_rows["Login"].iloc[0])
        
        # Determine participant name, safely ignoring 'nan'
        p_name_parts = []
        first_row = valid_rows.iloc[0]
        
        if "Firstname" in df.columns and pd.notnull(first_row["Firstname"]):
            fname = str(first_row["Firstname"]).strip()
            if fname.lower() != "nan" and fname:
                p_name_parts.append(fname)
                
        if "Lastname" in df.columns and pd.notnull(first_row["Lastname"]):
            lname = str(first_row["Lastname"]).strip()
            if lname.lower() != "nan" and lname:
                p_name_parts.append(lname)
                
        p_name = " ".join(p_name_parts)
        if not p_name: 
            p_name = final_login

        return ParsedDataset(
            participant=final_login,
            participant_name=p_name,
            exercise=str(unique_tasks[0]),
            course=str(valid_rows["Course Name"].iloc[0]) if "Course Name" in df.columns else "Unknown",
            source_file=Path(path).name,
            trials=trials,
            warnings=warnings
        )

    def _detect_header_row(self, path: str, sheet_name: str) -> Optional[int]:
        """Scan first 20 rows to find which one contains 'Login'."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb[sheet_name]
            for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                if "Login" in row:
                    return i
            return None
        except Exception:
            return None

    def _get_float(self, row: pd.Series, col: str) -> Optional[float]:
        if col in row.index and pd.notnull(row[col]):
            try:
                return float(row[col])
            except (ValueError, TypeError):
                return None
        return None

    def _get_int(self, row: pd.Series, col: str) -> Optional[int]:
        val = self._get_float(row, col)
        return int(val) if val is not None else None