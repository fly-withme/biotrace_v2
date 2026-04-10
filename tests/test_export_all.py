"""Tests for SessionExporter.export_all_sessions().

Run with:
    pytest tests/test_export_all.py -v
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from app.storage.database import DatabaseManager
from app.storage.session_repository import SessionRepository
from app.storage.calibration_repository import CalibrationRepository
from app.storage.export import SessionExporter


@pytest.fixture()
def db(tmp_path: Path) -> DatabaseManager:
    return DatabaseManager(db_path=str(tmp_path / "test.db"))


@pytest.fixture()
def two_sessions(db: DatabaseManager) -> tuple[int, int]:
    """Two completed sessions; returns (sid1, sid2)."""
    repo = SessionRepository(db)
    cal = CalibrationRepository(db)

    s1 = repo.create_session(datetime(2025, 1, 1, 9, 0, tzinfo=timezone.utc))
    repo.end_session(s1, datetime(2025, 1, 1, 9, 30, tzinfo=timezone.utc))
    cal.save_hrv_samples_bulk(s1, [(1.0, 850.0, 32.5, 70.6, 0.0)])

    s2 = repo.create_session(datetime(2025, 1, 2, 10, 0, tzinfo=timezone.utc))
    repo.end_session(s2, datetime(2025, 1, 2, 10, 45, tzinfo=timezone.utc))

    return s1, s2


class TestExportAllSessions:
    def test_creates_file(self, db: DatabaseManager, two_sessions, tmp_path: Path) -> None:
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        assert out.exists()

    def test_summary_sheet_exists(self, db: DatabaseManager, two_sessions, tmp_path: Path) -> None:
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames

    def test_summary_has_one_row_per_session(
        self, db: DatabaseManager, two_sessions, tmp_path: Path
    ) -> None:
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        # 1 header + 2 data rows
        assert ws.max_row == 3

    def test_summary_column_headers(
        self, db: DatabaseManager, two_sessions, tmp_path: Path
    ) -> None:
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Session ID" in headers
        assert "Date" in headers
        assert "Duration (s)" in headers

    def test_has_sheet_per_session(
        self, db: DatabaseManager, two_sessions, tmp_path: Path
    ) -> None:
        s1, s2 = two_sessions
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        wb = openpyxl.load_workbook(out)
        assert f"Session {s1}" in wb.sheetnames
        assert f"Session {s2}" in wb.sheetnames

    def test_session_sheet_has_measurements_columns(
        self, db: DatabaseManager, two_sessions, tmp_path: Path
    ) -> None:
        s1, _ = two_sessions
        out = tmp_path / "all.xlsx"
        SessionExporter(db).export_all_sessions(out)
        wb = openpyxl.load_workbook(out)
        ws = wb[f"Session {s1}"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Time" in headers
        assert "BPM" in headers
        assert "RMSSD" in headers

    def test_empty_database_writes_summary_only(
        self, db: DatabaseManager, tmp_path: Path
    ) -> None:
        out = tmp_path / "empty.xlsx"
        SessionExporter(db).export_all_sessions(out)
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert "Summary" in wb.sheetnames
        ws = wb["Summary"]
        # Header + 1 "no sessions" row
        assert ws.max_row == 2
