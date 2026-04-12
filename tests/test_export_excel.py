"""Unit tests for SessionExporter.export_excel().

Creates a real in-memory (tmp_path) SQLite database, inserts sample rows, and
verifies that the produced .xlsx file contains the expected sheets and data.

Run with:
    pytest tests/test_export_excel.py -v
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from app.storage.calibration_repository import CalibrationRepository
from app.storage.database import DatabaseManager
from app.storage.export import SessionExporter
from app.storage.session_repository import SessionRepository


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def db(tmp_path: Path) -> DatabaseManager:
    return DatabaseManager(db_path=str(tmp_path / "test.db"))


@pytest.fixture()
def session_id(db: DatabaseManager) -> int:
    repo = SessionRepository(db)
    start = datetime(2025, 1, 1, 10, 0, 0, tzinfo=timezone.utc)
    return repo.create_session(start)


@pytest.fixture()
def populated_session(db: DatabaseManager, session_id: int) -> int:
    """Insert HRV, pupil, and CLI rows for session_id; return session_id."""
    cal_repo = CalibrationRepository(db)

    # Two HRV beats at t=1.0 and t=2.0
    cal_repo.save_hrv_samples_bulk(
        session_id,
        [
            (1.0, 850.0, 32.5, 70.6, 0.0),
            (2.0, 820.0, 34.1, 73.2, 1.6),
        ],
    )
    # Two pupil samples at t=1.0 and t=3.0
    cal_repo.save_pupil_samples_bulk(
        session_id,
        [
            (1.0, 4.8, 4.9, 0.05),
            (3.0, 5.0, 5.1, 0.10),
        ],
    )
    # One CLI sample
    cal_repo.save_cli_samples_bulk(session_id, [(1.5, 0.42)])

    return session_id


# ---------------------------------------------------------------------------
# Basic file / sheet structure
# ---------------------------------------------------------------------------


class TestExportExcel:
    def test_creates_xlsx_file(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        assert out.exists()

    def test_xlsx_has_required_sheets(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        wb = openpyxl.load_workbook(out)
        assert "Session Info" in wb.sheetnames
        assert "Measurements" in wb.sheetnames

    def test_empty_session_still_creates_file(
        self, db: DatabaseManager, session_id: int, tmp_path: Path
    ) -> None:
        """A session with no samples should still produce a valid xlsx."""
        out = tmp_path / "export_empty.xlsx"
        SessionExporter(db).export_excel(session_id, out)
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert "Measurements" in wb.sheetnames


# ---------------------------------------------------------------------------
# Measurements sheet — columns and data
# ---------------------------------------------------------------------------


class TestMeasurementsSheet:
    EXPECTED_COLUMNS = [
        "Time",
        "BPM",
        "HRV",
        "RMSSD",
        "Delta RMSSD",
        "Pupil Diameter [px]",
        "Delta Pupil Dilation [PDI]",
    ]

    def _headers(self, wb: openpyxl.Workbook) -> list[str]:
        ws = wb["Measurements"]
        return [cell.value for cell in ws[1]]

    def _col(self, wb: openpyxl.Workbook, name: str) -> list:
        ws = wb["Measurements"]
        headers = [cell.value for cell in ws[1]]
        idx = headers.index(name) + 1
        return [ws.cell(row=r, column=idx).value for r in range(2, ws.max_row + 1)]

    def test_measurements_has_all_required_columns(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        headers = self._headers(openpyxl.load_workbook(out))
        for col in self.EXPECTED_COLUMNS:
            assert col in headers, f"Missing column: {col}"

    def test_measurements_columns_are_in_correct_order(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        headers = self._headers(openpyxl.load_workbook(out))
        assert headers == self.EXPECTED_COLUMNS

    def test_measurements_hrv_rows_present(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        """HRV samples contribute rows; outer merge means 2 HRV + 1 pupil-only = 3 rows."""
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        ws = openpyxl.load_workbook(out)["Measurements"]
        # 2 HRV rows (t=1.0, t=2.0) + 1 pupil-only row (t=3.0) = 3 data rows
        assert ws.max_row == 4  # 1 header + 3 data

    def test_measurements_bpm_values_are_correct(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        bpm = self._col(openpyxl.load_workbook(out), "BPM")
        # t=1.0 → 70.6, t=2.0 → 73.2, t=3.0 → None (pupil-only row)
        assert pytest.approx(bpm[0], abs=0.1) == 70.6
        assert pytest.approx(bpm[1], abs=0.1) == 73.2
        assert bpm[2] is None

    def test_measurements_hrv_column_is_rr_interval(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        hrv = self._col(openpyxl.load_workbook(out), "HRV")
        assert pytest.approx(hrv[0], abs=0.1) == 850.0
        assert pytest.approx(hrv[1], abs=0.1) == 820.0

    def test_measurements_pupil_diameter_is_average_of_left_right(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        pupil = self._col(openpyxl.load_workbook(out), "Pupil Diameter [px]")
        # t=1.0 → avg(4.8, 4.9) = 4.85; t=3.0 → avg(5.0, 5.1) = 5.05
        assert pytest.approx(pupil[0], abs=0.01) == 4.85
        assert pupil[1] is None  # HRV-only row (t=2.0)
        assert pytest.approx(pupil[2], abs=0.01) == 5.05

    def test_measurements_delta_pupil_is_pdi(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        delta = self._col(openpyxl.load_workbook(out), "Delta Pupil Dilation [PDI]")
        assert pytest.approx(delta[0], abs=0.001) == 0.05
        assert delta[1] is None
        assert pytest.approx(delta[2], abs=0.001) == 0.10

    def test_measurements_times_are_absolute_timestamps(
        self, db: DatabaseManager, populated_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(populated_session, out)
        times = self._col(openpyxl.load_workbook(out), "Time")
        assert times == [
            "2025-01-01, 10:00:01",
            "2025-01-01, 10:00:02",
            "2025-01-01, 10:00:03",
        ]

    def test_measurements_empty_session_has_only_header(
        self, db: DatabaseManager, session_id: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export_empty.xlsx"
        SessionExporter(db).export_excel(session_id, out)
        ws = openpyxl.load_workbook(out)["Measurements"]
        assert ws.max_row == 1  # header only


# ---------------------------------------------------------------------------
# Session Info enrichment — duration + summary metrics
# ---------------------------------------------------------------------------


@pytest.fixture()
def ended_session(db: DatabaseManager) -> int:
    """Session with known start/end times and 3 HRV samples; returns session_id."""
    repo = SessionRepository(db)
    cal_repo = CalibrationRepository(db)

    start = datetime(2025, 1, 1, 10, 0, 0, tzinfo=timezone.utc)
    end = datetime(2025, 1, 1, 10, 1, 5, tzinfo=timezone.utc)  # 65 s later

    sid = repo.create_session(start)
    repo.end_session(sid, end)

    cal_repo.save_hrv_samples_bulk(
        sid,
        [
            (1.0, 850.0, 32.5, 70.6, 0.0),
            (2.0, 820.0, 34.1, 73.2, 1.6),
            (3.0, 810.0, 35.0, 74.1, 0.9),
        ],
    )
    cal_repo.save_pupil_samples_bulk(
        sid,
        [
            (1.0, 4.8, 4.9, 0.10),
            (2.0, 5.0, 5.1, 0.20),
        ],
    )
    return sid


class TestSessionInfoEnrichment:
    def _info_dict(self, wb: openpyxl.Workbook) -> dict[str, object]:
        """Return {Field: Value} mapping from the Session Info sheet."""
        ws = wb["Session Info"]
        return {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }

    def test_session_info_contains_duration_field(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert "Duration (s)" in info

    def test_session_info_duration_value_is_correct(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert info["Duration (s)"] == 65

    def test_session_info_contains_average_metrics(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert pytest.approx(info["Average HRV (RMSSD)"], abs=0.001) == 33.8666666667
        assert pytest.approx(info["Average Pupil Dilation Change (PDI)"], abs=0.001) == 0.15

    def test_session_info_does_not_include_nasa_tlx(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert "NASA-TLX score" not in info

    def test_session_info_contains_hrv_sample_count_field(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert "HRV samples" in info

    def test_session_info_hrv_sample_count_is_correct(
        self, db: DatabaseManager, ended_session: int, tmp_path: Path
    ) -> None:
        out = tmp_path / "export.xlsx"
        SessionExporter(db).export_excel(ended_session, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert info["HRV samples"] == 3

    def test_session_info_duration_is_none_when_session_not_ended(
        self, db: DatabaseManager, session_id: int, tmp_path: Path
    ) -> None:
        """An open session (no ended_at) should export Duration (s) as None/empty."""
        out = tmp_path / "export_open.xlsx"
        SessionExporter(db).export_excel(session_id, out)
        info = self._info_dict(openpyxl.load_workbook(out))
        assert "Duration (s)" in info
        assert info["Duration (s)"] is None
